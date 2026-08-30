"""
Exporta la hoja 05_Matriz_Web del Directorio de Publicaciones a JSON, para que
scripts/seed-matriz.ts la use sin depender del paquete npm `xlsx` (tiene
vulnerabilidades altas sin parche). Se corre a mano, una vez, cuando la matriz
cambie -- no es parte del build ni del deploy.

Uso:
  python scripts/export-matriz.py "<ruta al Directorio_Publicaciones...xlsx>"
"""
import json
import re
import sys
import openpyxl

GRADO_A_NUM = {
    "Primer Grado": 1,
    "Segundo Grado": 2,
    "Tercer Grado": 3,
    "Cuarto Grado": 4,
    "Quinto Grado": 5,
    "Sexto Grado": 6,
}

RUTA_A_CODIGO = {
    "Ruta Base": "BASE",
    "Ruta Visual": "VISUAL",
    "Ruta Seguimiento": "SEGUIMIENTO",
    "Ruta Integral": "INTEGRAL",
}

FILE_EXT_RE = re.compile(r"\.(pdf|docx|xlsx)$", re.IGNORECASE)


def is_filename(value):
    return isinstance(value, str) and bool(FILE_EXT_RE.search(value.strip()))


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else None
    if not path:
        print("Falta la ruta al xlsx", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["05_Matriz_Web"]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True))

    header_idx = 3  # fila 4 (0-indexed 3), confirmado al inspeccionar el archivo
    header = rows[header_idx]
    data_rows = rows[header_idx + 1:]

    publicaciones = []
    archivos = {}  # nombreArchivo -> tipo (dedup)

    def track_file(nombre, tipo):
        if is_filename(nombre):
            archivos[nombre.strip()] = tipo

    for row in data_rows:
        if row[0] is None:
            continue
        rec = dict(zip(header, row))

        grado = GRADO_A_NUM[rec["Grado"]]
        ruta = RUTA_A_CODIGO[rec["Ruta"]]

        def iso(v):
            return v.isoformat() if v is not None else None

        publicaciones.append({
            "grado": grado,
            "ruta": ruta,
            "periodo": rec["Periodo"],
            "publicarEl": iso(rec["Publicar_el"]),
            "implementaInicio": iso(rec["Implementa_inicio"]),
            "implementaFin": iso(rec["Implementa_fin"]),
            "trimestre": rec["Trimestre"],
            "mesComercial": rec["Mes_comercial"],
            "compraQuincena": rec["Compra_Quincena"],
            "compraMes": rec["Compra_Mes"],
            "compraTrimestre": rec["Compra_Trimestre"],
            "compraCiclo": rec["Compra_Ciclo"],
            "planeacionArchivo": rec["Planeacion"] if is_filename(rec["Planeacion"]) else None,
            "fichasArchivo": rec["Fichas"] if is_filename(rec["Fichas"]) else None,
            "diapositivaS01Archivo": rec["Diapositiva_S01"] if is_filename(rec["Diapositiva_S01"]) else None,
            "diapositivaS02Archivo": rec["Diapositiva_S02"] if is_filename(rec["Diapositiva_S02"]) else None,
            "diapositivaS03Archivo": rec["Diapositiva_S03"] if is_filename(rec["Diapositiva_S03"]) else None,
            "seguimientoQuincenaArchivo": rec["Seguimiento_si_Quincena"] if is_filename(rec["Seguimiento_si_Quincena"]) else None,
            "seguimientoMesArchivo": rec["Seguimiento_si_Mes"] if is_filename(rec["Seguimiento_si_Mes"]) else None,
            "seguimientoTrimestreArchivo": rec["Seguimiento_si_Trimestre"] if is_filename(rec["Seguimiento_si_Trimestre"]) else None,
            "seguimientoCicloArchivo": rec["Seguimiento_si_Ciclo"] if is_filename(rec["Seguimiento_si_Ciclo"]) else None,
        })

        track_file(rec["Planeacion"], "planeacion")
        track_file(rec["Fichas"], "fichas")
        track_file(rec["Diapositiva_S01"], "diapositiva")
        track_file(rec["Diapositiva_S02"], "diapositiva")
        track_file(rec["Diapositiva_S03"], "diapositiva")
        track_file(rec["Seguimiento_si_Quincena"], "seguimiento")
        track_file(rec["Seguimiento_si_Mes"], "seguimiento")
        track_file(rec["Seguimiento_si_Trimestre"], "seguimiento")
        track_file(rec["Seguimiento_si_Ciclo"], "seguimiento")

    out = {
        "publicaciones": publicaciones,
        "archivos": [{"nombreArchivo": k, "tipo": v} for k, v in sorted(archivos.items())],
    }

    with open("data/matriz-web.json", "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f"OK: {len(publicaciones)} publicaciones, {len(archivos)} archivos unicos -> data/matriz-web.json")


if __name__ == "__main__":
    main()
